from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def cell_value(cell):
    return cell.value


def summarize_sheet(ws):
    merges = [str(rng) for rng in ws.merged_cells.ranges]

    columns = {}
    for key in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
        dim = ws.column_dimensions.get(key)
        if dim and dim.width is not None:
            columns[key] = dim.width

    row_heights = {}
    for idx, dim in ws.row_dimensions.items():
        if dim.height is not None:
            row_heights[int(idx)] = dim.height

    formulas = []
    styled_cells = []
    non_empty_rows = []

    for row in ws.iter_rows():
        row_values = []
        has_value = False
        for cell in row:
            value = cell_value(cell)
            if value is not None:
                has_value = True
                row_values.append(
                    {
                        "ref": cell.coordinate,
                        "value": value,
                    }
                )

            if isinstance(value, str) and value.startswith("="):
                formulas.append(
                    {
                        "ref": cell.coordinate,
                        "formula": value,
                    }
                )

            if cell.style_id:
                styled_cells.append(
                    {
                        "ref": cell.coordinate,
                        "style_id": cell.style_id,
                    }
                )

        if has_value:
            non_empty_rows.append(
                {
                    "row": row[0].row,
                    "cells": row_values,
                }
            )

    return {
        "title": ws.title,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "merged_cells_count": len(merges),
        "merged_cells_sample": merges[:50],
        "column_widths": columns,
        "row_heights_sample": dict(list(sorted(row_heights.items()))[:50]),
        "formulas": formulas[:100],
        "styled_cells_sample": styled_cells[:150],
        "non_empty_rows_sample": non_empty_rows[:120],
    }


def summarize_workbook(path: Path):
    wb = load_workbook(path)
    return {
        "file": str(path),
        "sheet_names": wb.sheetnames,
        "worksheets": [summarize_sheet(ws) for ws in wb.worksheets],
    }


def main():
    if len(sys.argv) < 2:
        raise SystemExit("Usage: inspect_xlsx.py <xlsx_path> [<xlsx_path> ...]")

    results = [summarize_workbook(Path(arg)) for arg in sys.argv[1:]]
    print(json.dumps(results, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
