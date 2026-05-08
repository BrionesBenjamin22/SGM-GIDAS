from copy import copy
from io import BytesIO
from pathlib import Path
from datetime import date
from typing import Iterable, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy.orm import joinedload, selectinload

from core.models.actividad_docencia import ActividadDocencia, InvestigadorActividadGrado
from core.models.articulo_divulgacion import ArticuloDivulgacion
from core.models.becas import Beca, Beca_Becario
from core.models.directivos import DirectivoGrupo
from core.models.distinciones import DistincionRecibida
from core.models.documentacion_autores import DocumentacionBibliografica
from core.models.equipamiento import Equipamiento
from core.models.erogacion import Erogacion
from core.models.grupo import GrupoInvestigacionUtn
from core.models.participacion_relevante import ParticipacionRelevante
from core.models.personal import Becario, Investigador, Personal
from core.models.proyecto_investigacion import ProyectoInvestigacion, InvestigadorProyecto, BecarioProyecto
from core.models.registro_patente import RegistrosPropiedad
from core.models.trabajo_reunion import TrabajoReunionCientifica
from core.models.trabajo_revista import TrabajosRevistasReferato
from core.models.transferencia_socio import TransferenciaSocioProductiva, AdoptanteTransferencia
from core.models.visita_grupo import VisitaAcademica
from core.models.programa_actividades import PlanificacionGrupo
from core.models.memorias import EstadoMemoria
from core.services.memoria_service import MemoriaService


class ExportService:
    MEMORIA_TEMPLATE_PATH = Path(__file__).resolve().parents[2] / "assets" / "DS2025 - UTN - PLANTILLA MEMORIAS.xlsx"
    TITLE_FILL = PatternFill(fill_type="solid", fgColor="FBE4D5")
    SECTION_FILL = PatternFill(fill_type="solid", fgColor="FBE4D5")
    SUBSECTION_FILL = PatternFill(fill_type="solid", fgColor="FFFF00")
    ACCENT_FILL = PatternFill(fill_type="solid", fgColor="FCE5CD")
    TOTAL_FILL = PatternFill(fill_type="solid", fgColor="D9EAD3")
    THIN_SIDE = Side(style="thin", color="000000")
    THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
    HEADER_FONT = Font(name="Calibri", size=11, bold=True)
    TITLE_FONT = Font(name="Calibri", size=15, bold=True)
    BODY_FONT = Font(name="Calibri", size=11)
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT_MIDDLE = Alignment(horizontal="left", vertical="center", wrap_text=True)
    LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
    COLUMN_WIDTHS = {
        "A": 23.14, "B": 40.57, "C": 36.43, "D": 23.14,
        "E": 27.43, "F": 34.14, "G": 65.14, "H": 74.86,
        "I": 81.43, "J": 29.0, "K": 13.0, "L": 13.0,
    }

    @staticmethod
    def _current_hours(entity):
        historial = getattr(entity, "historial_horas", None) or []
        activo = next((h for h in historial if h.fecha_fin is None), None)
        return activo.horas_semanales if activo else getattr(entity, "horas_semanales", None)

    @staticmethod
    def _active_grado_nombre(actividad: ActividadDocencia):
        historial = getattr(actividad, "investigadores_grado", None) or []
        grado_activo = next((h.grado_academico for h in historial if h.fecha_fin is None), None)
        return grado_activo.nombre if grado_activo else "-"

    @staticmethod
    def _money(value):
        return float(value or 0)

    @staticmethod
    def _join_names(items: Iterable, attr: str = "nombre_apellido", fallback: str = "-"):
        values = []
        for item in items:
            if getattr(item, "deleted_at", None) is None:
                value = getattr(item, attr, None)
                if value:
                    values.append(str(value))
        return ", ".join(values) if values else fallback

    @staticmethod
    def _safe_text(value, fallback="-"):
        if value is None:
            return fallback
        if isinstance(value, str):
            value = " ".join(value.split())
            return value if value else fallback
        return value

    @classmethod
    def _build_workbook(cls):
        wb = Workbook()
        ws = wb.active
        ws.title = "Memorias"
        ws.freeze_panes = "A4"
        for column, width in cls.COLUMN_WIDTHS.items():
            ws.column_dimensions[column].width = width
        return wb, ws

    @classmethod
    def _style_cell(cls, cell, *, font=None, fill=None, border=None, alignment=None, number_format=None):
        cell.font = font or cls.BODY_FONT
        if fill is not None:
            cell.fill = fill
        if border is not None:
            cell.border = border
        if alignment is not None:
            cell.alignment = alignment
        if number_format is not None:
            cell.number_format = number_format
        return cell

    @classmethod
    def _write_title(cls, ws, row: int, text: str):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
        cell = ws.cell(row=row, column=1, value=text)
        cls._style_cell(
            cell,
            font=cls.TITLE_FONT,
            fill=cls.TITLE_FILL,
            border=cls.THIN_BORDER,
            alignment=cls.CENTER,
        )
        ws.row_dimensions[row].height = 24
        return row + 2

    @classmethod
    def _write_section(cls, ws, row: int, text: str, span: int = 12):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        cell = ws.cell(row=row, column=1, value=text)
        cls._style_cell(
            cell,
            font=cls.HEADER_FONT,
            fill=cls.SECTION_FILL,
            border=cls.THIN_BORDER,
            alignment=cls.LEFT_TOP,
        )
        return row + 2

    @classmethod
    def _write_subsection(cls, ws, row: int, text: str, span: int = 8, accent: bool = False):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        cell = ws.cell(row=row, column=1, value=text)
        cls._style_cell(
            cell,
            font=cls.HEADER_FONT,
            fill=cls.ACCENT_FILL if accent else cls.SUBSECTION_FILL,
            border=cls.THIN_BORDER,
            alignment=cls.LEFT_TOP,
        )
        return row + 2

    @classmethod
    def _write_label_value(cls, ws, row: int, label: str, value, value_span: int = 5):
        label_cell = ws.cell(row=row, column=1, value=label)
        cls._style_cell(label_cell, font=cls.HEADER_FONT, border=cls.THIN_BORDER, alignment=cls.LEFT_TOP)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=1 + value_span)
        value_cell = ws.cell(row=row, column=2, value=value)
        cls._style_cell(value_cell, border=cls.THIN_BORDER, alignment=cls.LEFT_TOP)
        return row + 1

    @classmethod
    def _write_multiline_block(cls, ws, row: int, label: str, text: str, height: int = 72):
        row = cls._write_subsection(ws, row, label, span=8)
        ws.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=8)
        cell = ws.cell(row=row, column=1, value=cls._safe_text(text, fallback="Sin informacion registrada."))
        cls._style_cell(cell, border=cls.THIN_BORDER, alignment=cls.LEFT_TOP)
        ws.row_dimensions[row].height = height
        return row + 5

    @classmethod
    def _write_table(cls, ws, row: int, title: str, headers: Sequence[str], rows: Sequence[Sequence], *, accent: bool = False, merge_span: int | None = None, date_cols: set[int] | None = None, money_cols: set[int] | None = None):
        row = cls._write_subsection(ws, row, title, span=merge_span or max(len(headers), 4), accent=accent)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cls._style_cell(cell, font=cls.HEADER_FONT, border=cls.THIN_BORDER, alignment=cls.CENTER)
        ws.row_dimensions[row].height = 28
        row += 1
        if not rows:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(len(headers), 4))
            cell = ws.cell(row=row, column=1, value="No registra datos.")
            cls._style_cell(cell, border=cls.THIN_BORDER, alignment=cls.LEFT_TOP)
            return row + 3
        date_cols = date_cols or set()
        money_cols = money_cols or set()
        for row_data in rows:
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                number_format = None
                if col_idx in date_cols and value is not None:
                    number_format = "DD/MM/YYYY"
                elif col_idx in money_cols and value is not None:
                    number_format = '"$"#,##0.00'
                cls._style_cell(cell, border=cls.THIN_BORDER, alignment=cls.LEFT_TOP if col_idx != 1 else cls.CENTER, number_format=number_format)
            row += 1
        return row + 2

    @classmethod
    def _write_totals(cls, ws, row: int, label: str, values: Sequence[tuple[int, float]]):
        label_cell = ws.cell(row=row, column=1, value=label)
        cls._style_cell(
            label_cell,
            font=cls.HEADER_FONT,
            fill=cls.TOTAL_FILL,
            border=cls.THIN_BORDER,
            alignment=cls.LEFT_TOP,
        )
        for offset, (_, value) in enumerate(values, start=2):
            cell = ws.cell(row=row, column=offset, value=value)
            cls._style_cell(
                cell,
                font=cls.HEADER_FONT,
                fill=cls.TOTAL_FILL,
                border=cls.THIN_BORDER,
                alignment=cls.CENTER,
                number_format='"$"#,##0.00'
            )
        return row + 3

    @classmethod
    def _write_grouped_tables(
        cls,
        ws,
        row: int,
        base_index: str,
        base_title: str,
        groups: Sequence[tuple[str, Sequence[Sequence]]],
        headers: Sequence[str],
        *,
        accent: bool = False,
        merge_span: int | None = None,
        date_cols: set[int] | None = None,
        money_cols: set[int] | None = None,
    ):
        row = cls._write_subsection(
            ws,
            row,
            f"{base_index}.- {base_title}",
            span=merge_span or max(len(headers), 4),
            accent=accent,
        )

        subgroup_number = 1
        for group_title, group_rows in groups:
            if not group_rows:
                continue
            row = cls._write_table(
                ws,
                row,
                f"{base_index}.{subgroup_number}.- {group_title}",
                headers,
                group_rows,
                accent=accent,
                merge_span=merge_span,
                date_cols=date_cols,
                money_cols=money_cols,
            )
            subgroup_number += 1

        if subgroup_number == 1:
            row = cls._write_table(
                ws,
                row,
                f"{base_index}.1.- Sin registros",
                headers,
                [],
                accent=accent,
                merge_span=merge_span,
                date_cols=date_cols,
                money_cols=money_cols,
            )

        return row

    @staticmethod
    def _get_grupo(grupo_id: int | None):
        query = GrupoInvestigacionUtn.query.filter(GrupoInvestigacionUtn.deleted_at.is_(None))
        grupo = query.filter(GrupoInvestigacionUtn.id == grupo_id).first() if grupo_id is not None else query.order_by(GrupoInvestigacionUtn.id.asc()).first()
        if not grupo:
            raise ValueError("Grupo UTN no encontrado")
        return grupo

    @staticmethod
    def _get_directivos(grupo_id: int):
        return (
            DirectivoGrupo.query.options(joinedload(DirectivoGrupo.directivo), joinedload(DirectivoGrupo.cargo))
            .filter(DirectivoGrupo.id_grupo_utn == grupo_id, DirectivoGrupo.deleted_at.is_(None))
            .order_by(DirectivoGrupo.fecha_inicio.asc(), DirectivoGrupo.id.asc())
            .all()
        )

    @staticmethod
    def _get_investigadores(grupo_id: int):
        return (
            Investigador.query.options(joinedload(Investigador.categoria_utn), joinedload(Investigador.programa_incentivos), joinedload(Investigador.tipo_dedicacion), selectinload(Investigador.historial_horas))
            .filter(Investigador.grupo_utn_id == grupo_id, Investigador.deleted_at.is_(None))
            .order_by(Investigador.nombre_apellido.asc())
            .all()
        )

    @staticmethod
    def _get_personal(grupo_id: int):
        return (
            Personal.query.options(joinedload(Personal.tipo_personal), selectinload(Personal.historial_horas))
            .filter(Personal.grupo_utn_id == grupo_id, Personal.deleted_at.is_(None))
            .order_by(Personal.nombre_apellido.asc())
            .all()
        )

    @staticmethod
    def _get_becarios(grupo_id: int):
        return (
            Becario.query.options(joinedload(Becario.tipo_formacion), selectinload(Becario.historial_horas), selectinload(Becario.becas).joinedload(Beca_Becario.beca).joinedload(Beca.fuente_financiamiento))
            .filter(Becario.grupo_utn_id == grupo_id, Becario.deleted_at.is_(None))
            .order_by(Becario.nombre_apellido.asc())
            .all()
        )

    @staticmethod
    def _get_documentacion(grupo_id: int):
        return (
            DocumentacionBibliografica.query.options(selectinload(DocumentacionBibliografica.autores))
            .filter(DocumentacionBibliografica.grupo_id == grupo_id, DocumentacionBibliografica.deleted_at.is_(None))
            .order_by(DocumentacionBibliografica.anio.desc(), DocumentacionBibliografica.titulo.asc())
            .all()
        )

    @staticmethod
    def _get_actividades_docencia(grupo_id: int):
        return (
            ActividadDocencia.query.options(joinedload(ActividadDocencia.investigador), joinedload(ActividadDocencia.rol_actividad), selectinload(ActividadDocencia.investigadores_grado).joinedload(InvestigadorActividadGrado.grado_academico))
            .join(Investigador, ActividadDocencia.investigador_id == Investigador.id)
            .filter(ActividadDocencia.deleted_at.is_(None), Investigador.grupo_utn_id == grupo_id, Investigador.deleted_at.is_(None))
            .order_by(ActividadDocencia.fecha_inicio.desc(), ActividadDocencia.id.desc())
            .all()
        )

    @staticmethod
    def _get_articulos(grupo_id: int):
        return (
            ArticuloDivulgacion.query.filter(ArticuloDivulgacion.grupo_utn_id == grupo_id, ArticuloDivulgacion.deleted_at.is_(None))
            .order_by(ArticuloDivulgacion.fecha_publicacion.desc(), ArticuloDivulgacion.id.desc())
            .all()
        )

    @staticmethod
    def _get_becas(grupo_id: int):
        becas = (
            Beca.query.options(joinedload(Beca.fuente_financiamiento), selectinload(Beca.becarios).joinedload(Beca_Becario.becario))
            .filter(Beca.deleted_at.is_(None))
            .order_by(Beca.nombre_beca.asc())
            .all()
        )
        rows = []
        for beca in becas:
            for relacion in beca.becarios:
                becario = relacion.becario
                if relacion.deleted_at is None and becario is not None and becario.deleted_at is None and becario.grupo_utn_id == grupo_id:
                    rows.append((beca, relacion))
        return rows

    @staticmethod
    def _get_distinciones(grupo_id: int):
        return (
            DistincionRecibida.query.options(joinedload(DistincionRecibida.proyecto_investigacion))
            .join(ProyectoInvestigacion, DistincionRecibida.proyecto_investigacion_id == ProyectoInvestigacion.id)
            .filter(DistincionRecibida.deleted_at.is_(None), ProyectoInvestigacion.grupo_utn_id == grupo_id, ProyectoInvestigacion.deleted_at.is_(None))
            .order_by(DistincionRecibida.fecha.desc(), DistincionRecibida.id.desc())
            .all()
        )

    @staticmethod
    def _get_equipamiento(grupo_id: int):
        return (
            Equipamiento.query.filter(Equipamiento.grupo_utn_id == grupo_id, Equipamiento.deleted_at.is_(None))
            .order_by(Equipamiento.fecha_incorporacion.desc(), Equipamiento.id.desc())
            .all()
        )

    @staticmethod
    def _get_erogaciones(grupo_id: int):
        return (
            Erogacion.query.options(joinedload(Erogacion.tipo_erogacion), joinedload(Erogacion.fuente_financiamiento))
            .filter(Erogacion.grupo_utn_id == grupo_id, Erogacion.deleted_at.is_(None))
            .order_by(Erogacion.fecha.desc(), Erogacion.id.desc())
            .all()
        )

    @staticmethod
    def _get_participaciones(grupo_id: int):
        return (
            ParticipacionRelevante.query.options(joinedload(ParticipacionRelevante.investigador))
            .join(Investigador, ParticipacionRelevante.investigador_id == Investigador.id)
            .filter(ParticipacionRelevante.deleted_at.is_(None), Investigador.grupo_utn_id == grupo_id, Investigador.deleted_at.is_(None))
            .order_by(ParticipacionRelevante.fecha.desc(), ParticipacionRelevante.id.desc())
            .all()
        )

    @staticmethod
    def _get_registros(grupo_id: int):
        return (
            RegistrosPropiedad.query.options(joinedload(RegistrosPropiedad.tipo_registro))
            .filter(RegistrosPropiedad.grupo_utn_id == grupo_id, RegistrosPropiedad.deleted_at.is_(None))
            .order_by(RegistrosPropiedad.fecha_registro.desc(), RegistrosPropiedad.id.desc())
            .all()
        )

    @staticmethod
    def _get_trabajos_reunion(grupo_id: int):
        return (
            TrabajoReunionCientifica.query.options(joinedload(TrabajoReunionCientifica.tipo_reunion_cientifica), selectinload(TrabajoReunionCientifica.investigadores))
            .filter(TrabajoReunionCientifica.grupo_utn_id == grupo_id, TrabajoReunionCientifica.deleted_at.is_(None))
            .order_by(TrabajoReunionCientifica.fecha_inicio.desc(), TrabajoReunionCientifica.id.desc())
            .all()
        )

    @staticmethod
    def _get_trabajos_revista(grupo_id: int):
        return (
            TrabajosRevistasReferato.query.options(joinedload(TrabajosRevistasReferato.tipo_reunion), selectinload(TrabajosRevistasReferato.investigadores))
            .filter(TrabajosRevistasReferato.grupo_utn_id == grupo_id, TrabajosRevistasReferato.deleted_at.is_(None))
            .order_by(TrabajosRevistasReferato.fecha.desc(), TrabajosRevistasReferato.id.desc())
            .all()
        )

    @staticmethod
    def _get_transferencias(grupo_id: int):
        return (
            TransferenciaSocioProductiva.query.options(joinedload(TransferenciaSocioProductiva.tipo_contrato_transferencia), selectinload(TransferenciaSocioProductiva.participaciones).joinedload(AdoptanteTransferencia.adoptante))
            .filter(TransferenciaSocioProductiva.grupo_utn_id == grupo_id, TransferenciaSocioProductiva.deleted_at.is_(None))
            .order_by(TransferenciaSocioProductiva.fecha_inicio.desc(), TransferenciaSocioProductiva.id.desc())
            .all()
        )

    @staticmethod
    def _get_visitas(grupo_id: int):
        return (
            VisitaAcademica.query.options(joinedload(VisitaAcademica.tipo_visita))
            .filter(
                VisitaAcademica.grupo_utn_id == grupo_id,
                VisitaAcademica.deleted_at.is_(None),
            )
            .order_by(VisitaAcademica.fecha.desc(), VisitaAcademica.id.desc())
            .all()
        )

    @staticmethod
    def _get_planificaciones(grupo_id: int):
        return (
            PlanificacionGrupo.query.filter(
                PlanificacionGrupo.grupo_id == grupo_id,
                PlanificacionGrupo.deleted_at.is_(None),
            )
            .order_by(PlanificacionGrupo.anio.asc(), PlanificacionGrupo.id.asc())
            .all()
        )

    @staticmethod
    def _get_proyectos(grupo_id: int):
        return (
            ProyectoInvestigacion.query.options(joinedload(ProyectoInvestigacion.tipo_proyecto), joinedload(ProyectoInvestigacion.fuente_financiamiento), selectinload(ProyectoInvestigacion.participaciones_investigador).joinedload(InvestigadorProyecto.investigador), selectinload(ProyectoInvestigacion.participaciones_becario).joinedload(BecarioProyecto.becario), selectinload(ProyectoInvestigacion.distinciones))
            .filter(ProyectoInvestigacion.grupo_utn_id == grupo_id, ProyectoInvestigacion.deleted_at.is_(None))
            .order_by(ProyectoInvestigacion.fecha_inicio.desc(), ProyectoInvestigacion.id.desc())
            .all()
        )

    @staticmethod
    def _find_directivo_nombre(directivos, fragments: Sequence[str]):
        for participacion in directivos:
            if participacion.deleted_at is not None:
                continue
            cargo = participacion.cargo.nombre.lower() if participacion.cargo and participacion.cargo.nombre else ""
            if any(fragment in cargo for fragment in fragments):
                if participacion.directivo and participacion.directivo.deleted_at is None:
                    return participacion.directivo.nombre_apellido
        return "-"

    @staticmethod
    def _join_dict_names(items: Sequence[dict] | None, key: str, fallback: str = "-"):
        if not items:
            return fallback

        values = []
        for item in items:
            if not isinstance(item, dict):
                continue
            value = item.get(key)
            if value:
                values.append(str(value))

        return ", ".join(values) if values else fallback

    @staticmethod
    def _format_date(value, fallback="-"):
        if not value:
            return fallback
        if hasattr(value, "strftime"):
            return value.strftime("%d/%m/%Y")
        return str(value)

    @classmethod
    def _format_period(cls, fecha_inicio, fecha_fin):
        if not fecha_inicio and not fecha_fin:
            return "-"
        if fecha_inicio and fecha_fin:
            return f"{cls._format_date(fecha_inicio)} al {cls._format_date(fecha_fin)}"
        if fecha_inicio:
            return f"Desde {cls._format_date(fecha_inicio)}"
        return f"Hasta {cls._format_date(fecha_fin)}"

    @classmethod
    def _load_memoria_template(cls):
        return load_workbook(cls.MEMORIA_TEMPLATE_PATH)

    @staticmethod
    def _copy_cell_style(source, target):
        if source.has_style:
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.number_format = source.number_format
            target.protection = copy(source.protection)
        if source._style is not None:
            target._style = copy(source._style)

    @classmethod
    def _copy_row_style(cls, ws, source_row: int, target_row: int):
        for col_idx in range(1, ws.max_column + 1):
            source_cell = ws.cell(row=source_row, column=col_idx)
            target_cell = ws.cell(row=target_row, column=col_idx)
            cls._copy_cell_style(source_cell, target_cell)
        if ws.row_dimensions[source_row].height is not None:
            ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height

    @classmethod
    def _insert_styled_rows(cls, ws, insert_at: int, amount: int, source_row: int):
        if amount <= 0:
            return
        ws.insert_rows(insert_at, amount)
        for offset in range(amount):
            cls._copy_row_style(ws, source_row, insert_at + offset)

    @staticmethod
    def _resolve_write_cell(ws, row: int, column: str):
        col_idx = column_index_from_string(column)
        cell = ws.cell(row=row, column=col_idx)
        if not isinstance(cell, MergedCell):
            return cell

        for merged_range in ws.merged_cells.ranges:
            if (
                merged_range.min_row <= row <= merged_range.max_row
                and merged_range.min_col <= col_idx <= merged_range.max_col
            ):
                return ws.cell(row=merged_range.min_row, column=merged_range.min_col)

        return ws.cell(row=row, column=col_idx)

    @classmethod
    def _write_cell(cls, ws, row: int, column: str, value):
        target_cell = cls._resolve_write_cell(ws, row, column)
        target_cell.value = value
        return target_cell

    @classmethod
    def _normalize_generated_font(cls, cell, *, bold: bool = False):
        font = copy(cell.font) if cell.font is not None else copy(cls.BODY_FONT)
        font.name = font.name or cls.BODY_FONT.name
        font.size = font.size or cls.BODY_FONT.size
        font.bold = bold
        font.underline = None
        font.color = "FF000000"
        cell.font = font

    @classmethod
    def _write_body_cell(cls, ws, row: int, column: str, value):
        target_cell = cls._write_cell(ws, row, column, value)
        cls._normalize_generated_font(target_cell, bold=False)
        return target_cell

    @classmethod
    def _write_message_cell(cls, ws, row: int, column: str, value):
        target_cell = cls._write_cell(ws, row, column, value)
        cls._normalize_generated_font(target_cell, bold=False)
        return target_cell

    @classmethod
    def _clear_row_columns(cls, ws, row: int, columns: Sequence[str]):
        for column in columns:
            cls._write_cell(ws, row, column, None)

    @classmethod
    def _estimate_row_height(cls, values: Sequence, min_height: float = 18.0, chars_per_line: int = 48):
        text_lengths = [
            len(str(value).strip())
            for value in values
            if value not in (None, "")
        ]
        if not text_lengths:
            return min_height

        longest = max(text_lengths)
        lines = max(1, (longest // chars_per_line) + (1 if longest % chars_per_line else 0))
        return max(min_height, min(90.0, min_height * lines))

    @classmethod
    def _apply_row_height(cls, ws, row: int, values: Sequence, min_height: float = 18.0, chars_per_line: int = 48):
        ws.row_dimensions[row].height = cls._estimate_row_height(
            values,
            min_height=min_height,
            chars_per_line=chars_per_line,
        )

    @classmethod
    def _clear_merged_rows(cls, ws, rows: Sequence[int]):
        for row in rows:
            cls._set_merged_text(ws, row, "", normalize=False)
            ws.row_dimensions[row].height = 18.0

    @classmethod
    def _clear_rows_content(cls, ws, rows: Sequence[int]):
        for row in rows:
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col_idx)
                if isinstance(cell, MergedCell):
                    continue
                cell.value = None
                cell.border = Border()
                cell.fill = PatternFill(fill_type=None)
            ws.row_dimensions[row].height = 18.0

    @classmethod
    def _unmerge_rows(cls, ws, start_row: int, end_row: int):
        for merged_range in list(ws.merged_cells.ranges):
            if merged_range.min_row <= end_row and merged_range.max_row >= start_row:
                ws.unmerge_cells(str(merged_range))

    @classmethod
    def _write_single_merged_message(cls, ws, row: int, text: str, clear_rows: Sequence[int] | None = None):
        cls._set_merged_text(ws, row, text, normalize=True)
        cls._apply_row_height(ws, row, [text], min_height=18.0, chars_per_line=90)
        if clear_rows:
            cls._clear_merged_rows(ws, clear_rows)

    @classmethod
    def _write_section_table(
        cls,
        ws,
        title_row: int,
        data_start_row: int,
        data_end_row: int,
        title: str,
        headers: Sequence[str],
        rows: Sequence[Sequence],
        *,
        min_height: float = 18.0,
        chars_per_line: int = 42,
    ):
        last_col = len(headers)
        cls._unmerge_rows(ws, title_row, data_end_row)
        cls._clear_rows_content(ws, range(title_row, data_end_row + 1))

        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=last_col)
        title_cell = ws.cell(row=title_row, column=1, value=title)
        cls._style_cell(
            title_cell,
            font=cls.HEADER_FONT,
            fill=cls.ACCENT_FILL,
            border=cls.THIN_BORDER,
            alignment=cls.LEFT_TOP,
        )

        header_row = data_start_row
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cls._style_cell(
                cell,
                font=cls.HEADER_FONT,
                fill=cls.ACCENT_FILL,
                border=cls.THIN_BORDER,
                alignment=cls.CENTER,
            )
        ws.row_dimensions[header_row].height = 22.0

        body_rows = list(rows[: max(0, data_end_row - data_start_row)])
        current_row = header_row + 1
        if not body_rows:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=last_col)
            message_cell = ws.cell(row=current_row, column=1, value="No registra datos en esta memoria.")
            cls._style_cell(
                message_cell,
                font=cls.BODY_FONT,
                border=cls.THIN_BORDER,
                alignment=cls.LEFT_TOP,
            )
            cls._normalize_generated_font(message_cell, bold=False)
            cls._apply_row_height(ws, current_row, ["No registra datos en esta memoria."], min_height=min_height, chars_per_line=90)
            cls._clear_rows_content(ws, range(current_row + 1, data_end_row + 1))
            return

        for row_values in body_rows:
            for col_idx in range(1, last_col + 1):
                value = row_values[col_idx - 1] if col_idx - 1 < len(row_values) else ""
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cls._style_cell(
                    cell,
                    font=cls.BODY_FONT,
                    border=cls.THIN_BORDER,
                    alignment=cls.LEFT_MIDDLE if col_idx != 1 else cls.CENTER,
                )
                cls._normalize_generated_font(cell, bold=False)
            cls._apply_row_height(ws, current_row, row_values, min_height=min_height, chars_per_line=chars_per_line)
            current_row += 1

        cls._clear_rows_content(ws, range(current_row, data_end_row + 1))

    @classmethod
    def _write_total_row(
        cls,
        ws,
        row: int,
        label: str,
        total_columns: Sequence[tuple[int, float]],
        *,
        label_end_col: int,
    ):
        cls._unmerge_rows(ws, row, row)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=label_end_col)
        label_cell = ws.cell(row=row, column=1, value=label)
        cls._style_cell(
            label_cell,
            font=cls.HEADER_FONT,
            fill=cls.TOTAL_FILL,
            border=cls.THIN_BORDER,
            alignment=cls.LEFT_MIDDLE,
        )
        cls._normalize_generated_font(label_cell, bold=True)

        for col_idx, value in total_columns:
            total_cell = ws.cell(row=row, column=col_idx, value=value)
            cls._style_cell(
                total_cell,
                font=cls.HEADER_FONT,
                fill=cls.TOTAL_FILL,
                border=cls.THIN_BORDER,
                alignment=cls.CENTER,
                number_format='"$"#,##0.00',
            )
            cls._normalize_generated_font(total_cell, bold=True)

    @classmethod
    def _write_template_table(
        cls,
        ws,
        base_start: int,
        next_base: int,
        rows: Sequence[Sequence],
        columns: Sequence[str],
        offset: int,
        style_row: int | None = None,
        min_height: float = 18.0,
        chars_per_line: int = 48,
        empty_message: str = "No registra datos en esta memoria.",
    ):
        start_row = base_start + offset
        reserved_rows = max(next_base - base_start, 1)
        rows = list(rows)
        if len(rows) > reserved_rows:
            rows = rows[:reserved_rows]
        min_col_idx = min(column_index_from_string(column) for column in columns)
        max_col_idx = max(column_index_from_string(column) for column in columns)

        total_rows = reserved_rows
        for row_index in range(total_rows):
            current_row = start_row + row_index
            cls._clear_row_columns(ws, current_row, columns)
            if rows:
                if row_index < len(rows):
                    row_values = rows[row_index]
                    for column, value in zip(columns, row_values):
                        cls._write_body_cell(ws, current_row, column, value)
                    cls._apply_row_height(ws, current_row, row_values, min_height=min_height, chars_per_line=chars_per_line)
                else:
                    cls._clear_rows_content(ws, [current_row])
            elif row_index == 0:
                cls._unmerge_rows(ws, current_row, current_row)
                ws.merge_cells(start_row=current_row, start_column=min_col_idx, end_row=current_row, end_column=max_col_idx)
                message_cell = ws.cell(row=current_row, column=min_col_idx, value=empty_message)
                cls._style_cell(
                    message_cell,
                    font=cls.BODY_FONT,
                    border=cls.THIN_BORDER,
                    alignment=cls.LEFT_MIDDLE,
                )
                cls._normalize_generated_font(message_cell, bold=False)
                cls._apply_row_height(ws, current_row, [empty_message], min_height=min_height, chars_per_line=90)
            else:
                cls._clear_rows_content(ws, [current_row])
        return offset

    @classmethod
    def _set_merged_text(cls, ws, row: int, text: str, normalize: bool = False):
        target_cell = cls._write_cell(ws, row, "A", text)
        if normalize:
            cls._normalize_generated_font(target_cell, bold=False)

    @staticmethod
    def _infer_snapshot_grupo_id(snapshot_sources: dict):
        for key in (
            "investigadores",
            "becarios",
            "personal",
            "proyectos",
            "documentacion",
            "equipamiento",
            "erogaciones",
            "transferencias",
            "trabajos_reunion",
            "trabajos_revista",
            "articulos",
            "visitas",
            "registros",
        ):
            for item in snapshot_sources.get(key, []):
                grupo_id = item.get("grupo_utn_id") or item.get("grupo_id")
                if grupo_id:
                    return grupo_id
        return None

    @staticmethod
    def _summarize_text_items(items: Sequence[str], fallback="No registra datos en esta memoria."):
        clean_items = [str(item).strip() for item in items if item and str(item).strip()]
        if not clean_items:
            return fallback
        return " | ".join(clean_items)

    @classmethod
    def _build_text_lines(cls, rows: Sequence[dict], formatter, fallback="No registra datos en esta memoria."):
        lines = []
        for index, item in enumerate(rows, start=1):
            text = formatter(index, item)
            if text:
                lines.append(text)
        return cls._summarize_text_items(lines, fallback=fallback)

    @staticmethod
    def _clasificar_becario(tipo_formacion_nombre: str | None):
        tipo = (tipo_formacion_nombre or "").strip().lower()
        if "doctor" in tipo:
            return "doctorado"
        if "maestr" in tipo or "especial" in tipo:
            return "maestria"
        if "graduad" in tipo:
            return "graduado"
        if "pasant" in tipo:
            return "pasante"
        if "tesina" in tipo or "tesis" in tipo or "proyecto final" in tipo or "trabajo final" in tipo:
            return "tesina"
        return "alumno"

    @staticmethod
    def _clasificar_personal(tipo_personal_nombre: str | None):
        tipo = (tipo_personal_nombre or "").strip().lower()
        if "profes" in tipo:
            return "profesional"
        return "apoyo"

    @staticmethod
    def _clasificar_trabajo_reunion(tipo_reunion_nombre: str | None):
        tipo = (tipo_reunion_nombre or "").strip().lower()
        if "intern" in tipo:
            return "internacional"
        return "nacional"

    @staticmethod
    def _clasificar_transferencia(tipo_contrato_nombre: str | None):
        tipo = (tipo_contrato_nombre or "").strip().lower()
        if "tecnolog" in tipo:
            return "tecnologia"
        if "i+d" in tipo or "investig" in tipo:
            return "idi"
        if "conocimiento" in tipo:
            return "conocimientos"
        if "consult" in tipo or "asistencia" in tipo:
            return "asistencia"
        if "servicio" in tipo or "ensayo" in tipo or "laboratorio" in tipo or "supervisi" in tipo or "apoyo" in tipo:
            return "servicios"
        if "difusi" in tipo or "comunidad" in tipo:
            return "difusion"
        return "conocimientos"

    @staticmethod
    def _clasificar_erogacion(tipo_erogacion_nombre: str | None):
        tipo = (tipo_erogacion_nombre or "").strip().lower()
        if "capital" in tipo:
            return "capital"
        return "corriente"

    @staticmethod
    def _clasificar_registro(tipo_registro_nombre: str | None):
        tipo = (tipo_registro_nombre or "").strip().lower()
        if "industrial" in tipo:
            return "industrial"
        return "intelectual"

    @classmethod
    def _cleanup_template_placeholders(cls, ws):
        placeholder_exact = {
            "NOMBRE Y APELLIDO",
            "APELLIDO Y NOMBRE",
            "EJEMPLO",
            "999",
            "No aplica.",
        }
        placeholder_contains = (
            "Lorem ipsum",
            "NNNN/YYYY",
        )

        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if not isinstance(cell.value, str):
                    continue
                value = " ".join(cell.value.replace("\xa0", " ").split())
                if value in placeholder_exact:
                    cls._write_message_cell(ws, cell.row, cell.column_letter, "-")
                    continue
                if any(token in value for token in placeholder_contains):
                    cls._write_message_cell(ws, cell.row, cell.column_letter, "No registra datos en esta memoria.")

    @classmethod
    def _normalize_final_sheet(cls, ws):
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if cell.value in (None, ""):
                    continue

                font = copy(cell.font) if cell.font is not None else copy(cls.BODY_FONT)
                font.name = cls.BODY_FONT.name
                font.color = "FF000000"
                font.underline = None
                font.size = cls.TITLE_FONT.size if cell.row == 1 else cls.BODY_FONT.size
                if cell.row == 1:
                    font.bold = True
                elif cell.row == 2:
                    font.bold = True
                else:
                    font.bold = bool(font.bold)
                cell.font = font

                alignment = copy(cell.alignment) if cell.alignment is not None else copy(cls.LEFT_TOP)
                if alignment.wrap_text is None:
                    alignment.wrap_text = True
                if alignment.vertical is None:
                    alignment.vertical = "top"
                cell.alignment = alignment

    @staticmethod
    def _sheet_has_content(ws):
        for row in ws.iter_rows():
            for cell in row:
                if cell.value not in (None, ""):
                    return True
        return False

    @classmethod
    def _remove_empty_sheets(cls, wb):
        for ws in list(wb.worksheets):
            if ws.title != "Hoja1" and not cls._sheet_has_content(ws):
                wb.remove(ws)

    @classmethod
    def _build_memoria_snapshot_sources(cls, memoria_id: int, memoria_version_id: int):
        memoria = MemoriaService._get_memoria_or_404(memoria_id)
        version = MemoriaService._get_version_or_404(memoria_version_id)

        if version.memoria_id != memoria.id:
            raise ValueError("La version no pertenece a la memoria indicada")

        if version.estado != EstadoMemoria.CERRADA:
            raise ValueError("Solo puede exportarse una memoria con version cerrada")

        return {
            "memoria": memoria,
            "version": version,
            "investigadores": MemoriaService.get_investigadores_snapshot(memoria_id, memoria_version_id),
            "becarios": MemoriaService.get_becarios_snapshot(memoria_id, memoria_version_id),
            "personal": MemoriaService.get_personal_snapshot(memoria_id, memoria_version_id),
            "proyectos": MemoriaService.get_proyectos_snapshot(memoria_id, memoria_version_id),
            "actividades": MemoriaService.get_actividades_docencia_snapshot(memoria_id, memoria_version_id),
            "participaciones": MemoriaService.get_participaciones_relevantes_snapshot(memoria_id, memoria_version_id),
            "documentacion": MemoriaService.get_documentacion_snapshot(memoria_id, memoria_version_id),
            "equipamiento": MemoriaService.get_equipamiento_snapshot(memoria_id, memoria_version_id),
            "erogaciones": MemoriaService.get_erogaciones_snapshot(memoria_id, memoria_version_id),
            "transferencias": MemoriaService.get_transferencias_snapshot(memoria_id, memoria_version_id),
            "trabajos_reunion": MemoriaService.get_trabajos_reunion_snapshot(memoria_id, memoria_version_id),
            "trabajos_revista": MemoriaService.get_trabajos_revista_snapshot(memoria_id, memoria_version_id),
            "distinciones": MemoriaService.get_distinciones_snapshot(memoria_id, memoria_version_id),
            "registros": MemoriaService.get_registros_propiedad_snapshot(memoria_id, memoria_version_id),
            "articulos": MemoriaService.get_articulos_divulgacion_snapshot(memoria_id, memoria_version_id),
            "visitas": MemoriaService.get_visitas_snapshot(memoria_id, memoria_version_id),
        }

    @classmethod
    def generar_excel_memoria(cls, memoria_id: int, memoria_version_id: int):
        snapshot_sources = cls._build_memoria_snapshot_sources(
            memoria_id,
            memoria_version_id
        )

        memoria = snapshot_sources["memoria"]
        version = snapshot_sources["version"]
        grupo_id = cls._infer_snapshot_grupo_id(snapshot_sources)
        grupo = cls._get_grupo(grupo_id)
        directivos = cls._get_directivos(grupo.id)
        wb = cls._load_memoria_template()
        ws = wb["Hoja1"]

        anio_memoria = memoria.periodo_fin.year if memoria.periodo_fin else date.today().year
        cls._write_cell(ws, 1, "A", f"MEMORIAS {anio_memoria} DEL GRUPO UTN - {grupo.nombre_sigla_grupo}")
        cls._write_cell(ws, 2, "A", (
            f"Memoria {anio_memoria} - Version {version.numero_version} - "
            f"Periodo {cls._format_date(memoria.periodo_inicio)} al {cls._format_date(memoria.periodo_fin)}"
        ))

        cls._set_merged_text(ws, 5, f"1.1.- Facultad Regional: {grupo.nombre_unidad_academica}")
        cls._set_merged_text(ws, 6, f"1.2.- Nombre y Sigla: {grupo.nombre_sigla_grupo}")
        cls._set_merged_text(ws, 7, f"1.3.- Directora: {cls._find_directivo_nombre(directivos, ['director'])}")
        cls._set_merged_text(ws, 8, f"1.4.- Vicedirector/a: {cls._find_directivo_nombre(directivos, ['vicedirector', 'vice director'])}")
        cls._set_merged_text(ws, 9, f"1.5.- Direccion de Email: {grupo.mail}")

        offset = 0

        directivos_rows = [
            [idx, p.directivo.nombre_apellido, p.cargo.nombre if p.cargo else "-"]
            for idx, p in enumerate(directivos, start=1)
            if p.deleted_at is None and p.directivo and p.directivo.deleted_at is None
        ]
        offset = cls._write_template_table(ws, 12, 18, directivos_rows, ["A", "B", "D"], offset)

        cls._set_merged_text(ws, 18 + offset, "1.7.- Objetivos y desarrollo")
        cls._set_merged_text(
            ws,
            19 + offset,
            grupo.objetivo_desarrollo or "No registra datos institucionales adicionales para esta memoria."
        )
        cls._clear_merged_rows(
            ws,
            [row_number + offset for row_number in (20, 21, 22, 23, 24, 25, 28, 29, 30, 31, 33, 34, 35, 36, 37, 43)]
        )

        investigadores_rows = [
            [
                idx,
                item.get("nombre_apellido") or "-",
                item.get("categoria_utn_nombre") or "-",
                item.get("programa_incentivos_nombre") or "-",
                item.get("tipo_dedicacion_nombre") or "-",
                item.get("horas_semanales") or "-",
            ]
            for idx, item in enumerate(snapshot_sources["investigadores"], start=1)
        ]
        offset = cls._write_template_table(ws, 63, 72, investigadores_rows, ["A", "B", "C", "D", "E", "F"], offset)

        personal_profesional = [
            item for item in snapshot_sources["personal"]
            if cls._clasificar_personal(item.get("tipo_personal_nombre")) == "profesional"
        ]
        personal_apoyo = [
            item for item in snapshot_sources["personal"]
            if cls._clasificar_personal(item.get("tipo_personal_nombre")) != "profesional"
        ]
        offset = cls._write_template_table(
            ws,
            74,
            77,
            [[idx, item.get("nombre_apellido") or "-", item.get("horas_semanales") or "-"] for idx, item in enumerate(personal_profesional, start=1)],
            ["A", "B", "C"],
            offset
        )
        offset = cls._write_template_table(
            ws,
            79,
            81,
            [[idx, item.get("nombre_apellido") or "-", item.get("horas_semanales") or "-"] for idx, item in enumerate(personal_apoyo, start=1)],
            ["A", "B", "C"],
            offset
        )

        becarios_por_categoria = {
            "doctorado": [],
            "maestria": [],
            "graduado": [],
            "alumno": [],
            "pasante": [],
            "tesina": [],
        }
        for item in snapshot_sources["becarios"]:
            categoria = cls._clasificar_becario(item.get("tipo_formacion_nombre"))
            becarios_por_categoria[categoria].append(item)

        if becarios_por_categoria["doctorado"]:
            cls._set_merged_text(ws, 84 + offset, "")
        else:
            cls._set_merged_text(ws, 84 + offset, "No aplica.")
        offset = cls._write_template_table(
            ws,
            88,
            93,
            [
                [
                    idx,
                    item.get("nombre_apellido") or "-",
                    item.get("fuentes_financiamiento_beca") or "-",
                    item.get("horas_semanales") or "-",
                ]
                for idx, item in enumerate(becarios_por_categoria["maestria"], start=1)
            ],
            ["A", "B", "C", "D"],
            offset
        )
        offset = cls._write_template_table(
            ws,
            95,
            97,
            [
                [
                    idx,
                    item.get("nombre_apellido") or "-",
                    item.get("fuentes_financiamiento_beca") or "-",
                    item.get("horas_semanales") or "-",
                ]
                for idx, item in enumerate(becarios_por_categoria["graduado"], start=1)
            ],
            ["A", "B", "C", "D"],
            offset
        )
        offset = cls._write_template_table(
            ws,
            99,
            109,
            [
                [
                    idx,
                    item.get("nombre_apellido") or "-",
                    item.get("fuentes_financiamiento_beca") or "-",
                    item.get("horas_semanales") or "-",
                ]
                for idx, item in enumerate(becarios_por_categoria["alumno"], start=1)
            ],
            ["A", "B", "C", "D"],
            offset
        )
        offset = cls._write_template_table(
            ws,
            111,
            117,
            [
                [
                    idx,
                    item.get("nombre_apellido") or "-",
                    item.get("fuentes_financiamiento_beca") or "-",
                    item.get("horas_semanales") or "-",
                ]
                for idx, item in enumerate(becarios_por_categoria["pasante"], start=1)
            ],
            ["A", "B", "C", "D"],
            offset
        )
        offset = cls._write_template_table(
            ws,
            120,
            126,
            [
                [
                    idx,
                    item.get("nombre_apellido") or "-",
                    item.get("fuentes_financiamiento_beca") or "Sin financiamiento",
                    item.get("horas_semanales") or "-",
                ]
                for idx, item in enumerate(becarios_por_categoria["tesina"], start=1)
            ],
            ["A", "B", "C", "D"],
            offset
        )

        equipamiento_rows = [
            [
                idx,
                item.get("denominacion") or "-",
                cls._format_date(item.get("fecha_incorporacion")),
                cls._money(item.get("monto_invertido")),
                item.get("descripcion_breve") or "-",
            ]
            for idx, item in enumerate(snapshot_sources["equipamiento"], start=1)
        ]
        offset = cls._write_template_table(
            ws,
            128,
            131,
            equipamiento_rows,
            ["A", "B", "C", "D", "E"],
            offset,
            min_height=18.0,
            chars_per_line=55,
        )

        documentacion_rows = [
            [
                idx,
                item.get("titulo") or "-",
                cls._join_dict_names(item.get("autores"), "nombre_apellido"),
                item.get("editorial") or "-",
                item.get("anio") or "-",
                cls._format_date(item.get("fecha")),
            ]
            for idx, item in enumerate(snapshot_sources["documentacion"], start=1)
        ]
        offset = cls._write_template_table(
            ws,
            133,
            136,
            documentacion_rows,
            ["A", "B", "C", "D", "E", "F"],
            offset,
            min_height=18.0,
            chars_per_line=60,
        )

        proyectos_rows = [
            [
                idx,
                item.get("tipo_proyecto_nombre") or "-",
                item.get("codigo_proyecto") or "-",
                cls._format_period(item.get("fecha_inicio"), item.get("fecha_fin")),
                cls._money(item.get("monto_destinado")),
                item.get("nombre_proyecto") or "-",
                item.get("descripcion_proyecto") or "-",
                "Sin registro especifico en snapshot.",
                item.get("dificultades_proyecto") or "-",
                item.get("fuente_financiamiento_nombre") or "-",
            ]
            for idx, item in enumerate(snapshot_sources["proyectos"], start=1)
        ]
        cls._write_section_table(
            ws,
            138 + offset,
            139 + offset,
            143 + offset,
            "5.- Proyectos en curso",
            ["Nro.", "Tipo", "Codigo", "Periodo", "Monto destinado", "Proyecto", "Descripcion", "Logros", "Dificultades", "Fuente"],
            proyectos_rows,
            chars_per_line=28,
        )
        cls._write_total_row(
            ws,
            143 + offset,
            "Total monto proyectos",
            [(5, sum(cls._money(item.get("monto_destinado")) for item in snapshot_sources["proyectos"]))],
            label_end_col=4,
        )

        distinciones_rows = [
            [
                idx,
                item.get("descripcion") or "-",
                cls._format_date(item.get("fecha")),
                item.get("proyecto_nombre") or item.get("proyecto_investigacion_nombre") or "-",
            ]
            for idx, item in enumerate(snapshot_sources["distinciones"], start=1)
        ]
        cls._write_section_table(
            ws,
            144 + offset,
            145 + offset,
            148 + offset,
            "6.1.- Distinciones recibidas",
            ["Nro.", "Distincion / Premio", "Fecha", "Observaciones"],
            distinciones_rows,
            chars_per_line=38,
        )

        participaciones_rows = [
            [
                idx,
                item.get("nombre_evento") or "-",
                item.get("forma_participacion") or "-",
            ]
            for idx, item in enumerate(snapshot_sources["participaciones"], start=1)
        ]
        cls._write_section_table(
            ws,
            149 + offset,
            150 + offset,
            154 + offset,
            "6.2.- Participaciones",
            ["Nro.", "Evento", "Rol / Participacion"],
            participaciones_rows,
            chars_per_line=42,
        )

        visitas_rows = [
            [
                idx,
                item.get("procedencia") or "-",
                item.get("razon") or "-",
                item.get("tipo_visita_nombre") or "-",
            ]
            for idx, item in enumerate(snapshot_sources["visitas"], start=1)
        ]
        cls._write_section_table(
            ws,
            155 + offset,
            156 + offset,
            160 + offset,
            "6.3.- Visitantes del pais y del extranjero",
            ["Nro.", "Institucion / Procedencia", "Nombre o descripcion", "Tipo"],
            visitas_rows,
            chars_per_line=36,
        )

        cls._clear_rows_content(ws, [row_number + offset for row_number in range(161, 185)])

        reuniones_nacionales = []
        reuniones_internacionales = []
        for idx, item in enumerate(snapshot_sources["trabajos_reunion"], start=1):
            row_data = [
                idx,
                item.get("nombre_reunion") or "-",
                item.get("procedencia") or "-",
                cls._format_date(item.get("fecha_inicio")),
                item.get("investigadores_participantes") or "-",
                item.get("titulo_trabajo") or "-",
            ]
            if cls._clasificar_trabajo_reunion(item.get("tipo_reunion_nombre")) == "internacional":
                reuniones_internacionales.append(row_data)
            else:
                reuniones_nacionales.append(row_data)

        cls._write_section_table(
            ws,
            187 + offset,
            188 + offset,
            201 + offset,
            "7.1.- Reunion Cientifica Nacional con Referato",
            ["Nro.", "Nombre reunion", "Ciudad / Pais", "Fecha inicio", "Expositor", "Titulo trabajo"],
            reuniones_nacionales,
            chars_per_line=32,
        )
        cls._write_section_table(
            ws,
            202 + offset,
            203 + offset,
            205 + offset,
            "7.2.- Reunion Cientifica Internacional",
            ["Nro.", "Nombre reunion", "Ciudad / Pais", "Fecha inicio", "Expositor", "Titulo trabajo"],
            reuniones_internacionales,
            chars_per_line=32,
        )

        trabajos_revista_rows = [
            [
                idx,
                item.get("nombre_revista") or "-",
                item.get("pais") or "-",
                item.get("editorial") or "-",
                item.get("issn") or "-",
                item.get("titulo_trabajo") or "-",
            ]
            for idx, item in enumerate(snapshot_sources["trabajos_revista"], start=1)
        ]
        cls._write_section_table(
            ws,
            207 + offset,
            208 + offset,
            214 + offset,
            "8.1.- Trabajos publicados en revistas con referato",
            ["Nro.", "Revista", "Pais", "Editorial", "ISSN", "Titulo trabajo"],
            trabajos_revista_rows,
            chars_per_line=34,
        )

        articulos_rows = [
            [
                idx,
                item.get("titulo") or "-",
                item.get("descripcion") or "-",
                cls._format_date(item.get("fecha_publicacion")),
            ]
            for idx, item in enumerate(snapshot_sources["articulos"], start=1)
        ]
        cls._write_section_table(
            ws,
            221 + offset,
            222 + offset,
            228 + offset,
            "8.4.- Articulos de divulgacion, informes y memorias tecnicas",
            ["Nro.", "Titulo", "Descripcion", "Fecha"],
            articulos_rows,
            chars_per_line=42,
        )
        cls._clear_rows_content(ws, [215 + offset, 216 + offset, 229 + offset])

        registros_intelectuales = [
            item for item in snapshot_sources["registros"]
            if cls._clasificar_registro(item.get("tipo_registro_nombre")) == "intelectual"
        ]
        registros_industriales = [
            item for item in snapshot_sources["registros"]
            if cls._clasificar_registro(item.get("tipo_registro_nombre")) == "industrial"
        ]
        registros_intelectuales_rows = [
            [
                idx,
                item.get("nombre_articulo") or "-",
                item.get("organismo_registrante") or "-",
                cls._format_date(item.get("fecha_registro")),
            ]
            for idx, item in enumerate(registros_intelectuales, start=1)
        ]
        cls._write_section_table(
            ws,
            231 + offset,
            232 + offset,
            234 + offset,
            "9.1.- Registro de Propiedad Intelectual",
            ["Nro.", "Registro", "Organismo", "Fecha"],
            registros_intelectuales_rows,
            chars_per_line=34,
        )
        registros_industriales_rows = [
            [
                idx,
                item.get("nombre_articulo") or "-",
                item.get("organismo_registrante") or "-",
                cls._format_date(item.get("fecha_registro")),
            ]
            for idx, item in enumerate(registros_industriales, start=1)
        ]
        cls._write_section_table(
            ws,
            235 + offset,
            236 + offset,
            238 + offset,
            "9.2.- Registro de Propiedad Industrial",
            ["Nro.", "Registro", "Organismo", "Fecha"],
            registros_industriales_rows,
            chars_per_line=34,
        )

        actividades_rows = [
            [
                idx,
                item.get("investigador_nombre") or "-",
                item.get("grado_academico_nombre") or "-",
                f"{item.get('curso') or '-'} - {item.get('institucion') or '-'}",
            ]
            for idx, item in enumerate(snapshot_sources["actividades"], start=1)
        ]
        cls._write_section_table(
            ws,
            239 + offset,
            240 + offset,
            247 + offset,
            "III.- Actividades en docencia",
            ["Nro.", "Investigador", "Grado", "Actividades y catedras de posgrado"],
            actividades_rows,
            chars_per_line=30,
        )
        cls._clear_rows_content(ws, [248 + offset])

        transferencias_por_tipo = {
            "tecnologia": [],
            "idi": [],
            "conocimientos": [],
            "asistencia": [],
            "servicios": [],
            "difusion": [],
        }
        for item in snapshot_sources["transferencias"]:
            transferencias_por_tipo[cls._clasificar_transferencia(item.get("tipo_contrato_nombre"))].append(item)

        def transfer_rows(items):
            return [
                [
                    idx,
                    item.get("denominacion") or "-",
                    cls._join_dict_names(item.get("adoptantes"), "adoptante_nombre"),
                    item.get("demandante") or "-",
                    cls._money(item.get("monto")),
                    item.get("descripcion_actividad") or "-",
                ]
                for idx, item in enumerate(items, start=1)
            ]

        transfer_headers = ["Nro.", "Denominacion", "Adoptante", "Demandante", "Monto comprometido", "Breve descripcion"]
        cls._write_section_table(ws, 250 + offset, 251 + offset, 255 + offset, "10.1.- Contrato de transferencia de tecnologia", transfer_headers, transfer_rows(transferencias_por_tipo["tecnologia"]), chars_per_line=30)
        cls._write_section_table(ws, 256 + offset, 257 + offset, 260 + offset, "10.2.- Contrato de I+D+i", transfer_headers, transfer_rows(transferencias_por_tipo["idi"]), chars_per_line=30)
        cls._write_section_table(ws, 261 + offset, 262 + offset, 264 + offset, "10.3.- Contrato/Acuerdo de Transferencia de conocimientos", transfer_headers, transfer_rows(transferencias_por_tipo["conocimientos"]), chars_per_line=30)
        cls._write_section_table(ws, 265 + offset, 266 + offset, 269 + offset, "10.4.- Contrato de Asistencia Tecnica o Consultoria", transfer_headers, transfer_rows(transferencias_por_tipo["asistencia"]), chars_per_line=30)
        cls._write_section_table(ws, 270 + offset, 271 + offset, 273 + offset, "10.5.- Servicios tecnicos/de apoyo/supervision y/o ensayos de laboratorio", transfer_headers, transfer_rows(transferencias_por_tipo["servicios"]), chars_per_line=30)
        cls._write_section_table(ws, 274 + offset, 275 + offset, 278 + offset, "10.6.- Difusion a la comunidad academica y en general", transfer_headers, transfer_rows(transferencias_por_tipo["difusion"]), chars_per_line=30)
        cls._write_total_row(
            ws,
            278 + offset,
            "Total monto transferencias",
            [(5, sum(cls._money(item.get("monto")) for item in snapshot_sources["transferencias"]))],
            label_end_col=4,
        )

        erogaciones_corrientes = [
            item for item in snapshot_sources["erogaciones"]
            if cls._clasificar_erogacion(item.get("tipo_erogacion_nombre")) == "corriente"
        ]
        erogaciones_capital = [
            item for item in snapshot_sources["erogaciones"]
            if cls._clasificar_erogacion(item.get("tipo_erogacion_nombre")) == "capital"
        ]
        erogaciones_corrientes_rows = [
            [
                idx,
                item.get("fuente_financiamiento_nombre") or "-",
                cls._money(item.get("ingresos")),
                cls._money(item.get("egresos")),
                cls._money(item.get("ingresos")) - cls._money(item.get("egresos")),
            ]
            for idx, item in enumerate(erogaciones_corrientes, start=1)
        ]
        cls._write_section_table(
            ws,
            281 + offset,
            282 + offset,
            289 + offset,
            "11.1.- Erogaciones Corrientes",
            ["Nro.", "Fuente de financiamiento", "Ingresos", "Egresos", "Saldo resultante"],
            erogaciones_corrientes_rows,
            chars_per_line=30,
        )
        cls._write_total_row(
            ws,
            289 + offset,
            "Totales erogaciones",
            [
                (3, sum(cls._money(item.get("ingresos")) for item in erogaciones_corrientes)),
                (4, sum(cls._money(item.get("egresos")) for item in erogaciones_corrientes)),
                (5, sum(cls._money(item.get("ingresos")) - cls._money(item.get("egresos")) for item in erogaciones_corrientes)),
            ],
            label_end_col=2,
        )
        erogaciones_capital_rows = [
            [
                idx,
                item.get("fuente_financiamiento_nombre") or "-",
                cls._money(item.get("ingresos")),
                cls._money(item.get("egresos")),
                cls._money(item.get("ingresos")) - cls._money(item.get("egresos")),
            ]
            for idx, item in enumerate(erogaciones_capital, start=1)
        ]
        cls._write_section_table(
            ws,
            290 + offset,
            291 + offset,
            294 + offset,
            "11.2.- Erogaciones de Capital",
            ["Nro.", "Fuente de financiamiento", "Ingresos", "Egresos", "Saldo resultante"],
            erogaciones_capital_rows,
            chars_per_line=30,
        )
        cls._write_total_row(
            ws,
            294 + offset,
            "Totales erogaciones",
            [
                (3, sum(cls._money(item.get("ingresos")) for item in erogaciones_capital)),
                (4, sum(cls._money(item.get("egresos")) for item in erogaciones_capital)),
                (5, sum(cls._money(item.get("ingresos")) - cls._money(item.get("egresos")) for item in erogaciones_capital)),
            ],
            label_end_col=2,
        )

        cls._write_cell(ws, 295 + offset, "A", f"VI - PROGRAMA DE ACTIVIDADES para {anio_memoria + 1}")
        cls._set_merged_text(ws, 296 + offset, "No registra datos de planificacion en snapshots de memoria.", normalize=True)
        cls._apply_row_height(
            ws,
            296 + offset,
            ["No registra datos de planificacion en snapshots de memoria."],
            min_height=18.0,
            chars_per_line=90,
        )

        cls._cleanup_template_placeholders(ws)
        cls._normalize_final_sheet(ws)
        cls._remove_empty_sheets(wb)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @classmethod
    def generar_excel_grupo(cls, grupo_id: int | None = 1):
        grupo = cls._get_grupo(grupo_id)
        wb, ws = cls._build_workbook()

        directivos = cls._get_directivos(grupo.id)
        investigadores = cls._get_investigadores(grupo.id)
        personal = cls._get_personal(grupo.id)
        becarios = cls._get_becarios(grupo.id)
        documentacion = cls._get_documentacion(grupo.id)
        actividades = cls._get_actividades_docencia(grupo.id)
        articulos = cls._get_articulos(grupo.id)
        becas = cls._get_becas(grupo.id)
        distinciones = cls._get_distinciones(grupo.id)
        equipamiento = cls._get_equipamiento(grupo.id)
        erogaciones = cls._get_erogaciones(grupo.id)
        participaciones = cls._get_participaciones(grupo.id)
        registros = cls._get_registros(grupo.id)
        trabajos_reunion = cls._get_trabajos_reunion(grupo.id)
        trabajos_revista = cls._get_trabajos_revista(grupo.id)
        transferencias = cls._get_transferencias(grupo.id)
        visitas = cls._get_visitas(grupo.id)
        planificaciones = cls._get_planificaciones(grupo.id)
        proyectos = cls._get_proyectos(grupo.id)

        row = 1
        row = cls._write_title(ws, row, f"MEMORIAS {date.today().year} DEL GRUPO UTN - {grupo.nombre_sigla_grupo}")
        row = cls._write_section(ws, row, "I.- ADMINISTRACION")
        row = cls._write_subsection(ws, row, "1.- INDIVIDUALIZACION DEL GRUPO UTN", span=8)
        row = cls._write_label_value(ws, row, "1.1.- Facultad Regional", grupo.nombre_unidad_academica)
        row = cls._write_label_value(ws, row, "1.2.- Nombre y Sigla", grupo.nombre_sigla_grupo)
        row = cls._write_label_value(ws, row, "1.3.- Director/a", cls._find_directivo_nombre(directivos, ["director"]))
        row = cls._write_label_value(ws, row, "1.4.- Vicedirector/a", cls._find_directivo_nombre(directivos, ["vicedirector", "vice director"]))
        row = cls._write_label_value(ws, row, "1.5.- Direccion de Email", grupo.mail)
        row += 1
        directivos_rows = [[idx, p.directivo.nombre_apellido if p.directivo and p.directivo.deleted_at is None else "-", p.cargo.nombre if p.cargo else "-", p.fecha_inicio, p.fecha_fin] for idx, p in enumerate(directivos, start=1) if p.directivo and p.directivo.deleted_at is None]
        row = cls._write_table(ws, row, "1.6.- Autoridades y cargos de gestion del grupo", ["Nro.", "Apellido y nombre", "Cargo desempenado", "Fecha de inicio", "Fecha de finalizacion"], directivos_rows, merge_span=8, date_cols={4, 5})
        row = cls._write_multiline_block(ws, row, "1.7.- Objetivos y desarrollo", grupo.objetivo_desarrollo)
        row = cls._write_subsection(ws, row, "2.- PERSONAL", span=10)
        investigadores_rows = [[idx, inv.nombre_apellido, inv.categoria_utn.nombre if inv.categoria_utn else "-", inv.programa_incentivos.nombre if inv.programa_incentivos else "-", inv.tipo_dedicacion.nombre if inv.tipo_dedicacion else "-", cls._current_hours(inv)] for idx, inv in enumerate(investigadores, start=1)]
        row = cls._write_table(ws, row, "2.1.- Investigadores", ["Nro.", "Apellido y nombre", "Categoria UTN", "Programa de incentivos", "Tipo de dedicacion", "Carga horaria semanal"], investigadores_rows, merge_span=8)
        personal_rows = [[idx, persona.nombre_apellido, persona.tipo_personal.nombre if persona.tipo_personal else "-", cls._current_hours(persona)] for idx, persona in enumerate(personal, start=1)]
        row = cls._write_table(ws, row, "2.2.- Personal tecnico, administrativo y de apoyo", ["Nro.", "Apellido y nombre", "Categoria de personal", "Carga horaria semanal"], personal_rows, merge_span=8)

        becarios_rows = []
        for idx, becario in enumerate(becarios, start=1):
            relaciones_activas = [rel for rel in becario.becas if rel.deleted_at is None and rel.beca and rel.beca.deleted_at is None]
            nombres_becas = ", ".join(sorted({rel.beca.nombre_beca for rel in relaciones_activas})) or "-"
            fuentes = ", ".join(sorted({rel.beca.fuente_financiamiento.nombre for rel in relaciones_activas if rel.beca.fuente_financiamiento is not None})) or "-"
            becarios_rows.append([idx, becario.nombre_apellido, becario.tipo_formacion.nombre if becario.tipo_formacion else "-", nombres_becas, fuentes, cls._current_hours(becario)])
        row = cls._write_table(ws, row, "2.3.- Becarios y personal en formacion", ["Nro.", "Apellido y nombre", "Tipo de formacion", "Becas asociadas", "Fuente de financiamiento", "Carga horaria semanal"], becarios_rows, merge_span=8)
        equipamiento_rows = [[idx, item.denominacion, item.descripcion_breve, item.fecha_incorporacion, cls._money(item.monto_invertido)] for idx, item in enumerate(equipamiento, start=1)]
        row = cls._write_table(ws, row, "3.- EQUIPAMIENTO DEL GRUPO", ["Nro.", "Denominacion del equipamiento", "Descripcion breve", "Fecha de incorporacion", "Monto invertido"], equipamiento_rows, accent=True, merge_span=8, date_cols={4}, money_cols={5})
        docs_rows = [[idx, doc.titulo, ", ".join(a.nombre_apellido for a in doc.autores if not hasattr(a, "deleted_at") or a.deleted_at is None) or "-", doc.editorial, doc.anio] for idx, doc in enumerate(documentacion, start=1)]
        row = cls._write_table(ws, row, "4.- DOCUMENTACION Y BIBLIOTECA", ["Nro.", "Titulo de la obra", "Autores", "Editorial", "Anio de publicacion"], docs_rows, merge_span=8)

        row = cls._write_section(ws, row, "II.- ACTIVIDADES DE I+D+I")
        proyectos_rows = []
        total_proyectos = 0.0
        for idx, proyecto in enumerate(proyectos, start=1):
            monto = cls._money(proyecto.monto_destinado)
            total_proyectos += monto
            investigadores_activos = [rel.investigador for rel in proyecto.participaciones_investigador if rel.deleted_at is None and rel.investigador is not None and rel.investigador.deleted_at is None]
            becarios_activos = [rel.becario for rel in proyecto.participaciones_becario if rel.deleted_at is None and rel.becario is not None and rel.becario.deleted_at is None]
            distinciones_texto = "; ".join(dist.descripcion for dist in proyecto.distinciones if dist.deleted_at is None) or "-"
            proyectos_rows.append([idx, proyecto.codigo_proyecto, proyecto.nombre_proyecto, proyecto.tipo_proyecto.nombre if proyecto.tipo_proyecto else "-", proyecto.fuente_financiamiento.nombre if proyecto.fuente_financiamiento else "-", monto, proyecto.fecha_inicio, proyecto.fecha_fin, ", ".join(inv.nombre_apellido for inv in investigadores_activos) or "-", ", ".join(bec.nombre_apellido for bec in becarios_activos) or "-", distinciones_texto])
        row = cls._write_table(ws, row, "5.- PROYECTOS DE INVESTIGACION", ["Nro.", "Codigo del proyecto", "Denominacion del proyecto", "Tipo de proyecto", "Fuente de financiamiento", "Monto destinado", "Fecha de inicio", "Fecha de finalizacion", "Investigadores vinculados", "Becarios vinculados", "Distinciones asociadas"], proyectos_rows, accent=True, merge_span=12, date_cols={7, 8}, money_cols={6})
        row = cls._write_totals(ws, row, "Total monto proyectos", [(6, total_proyectos)])
        participaciones_rows = [[idx, participacion.nombre_evento, participacion.forma_participacion, participacion.fecha, participacion.investigador.nombre_apellido if participacion.investigador else "-"] for idx, participacion in enumerate(participaciones, start=1)]
        row = cls._write_subsection(ws, row, "6.- RECONOCIMIENTOS, PARTICIPACIONES Y VISITAS ACADEMICAS", span=10)
        distinciones_rows = [[idx, distincion.fecha, distincion.descripcion, distincion.proyecto_investigacion.nombre_proyecto if distincion.proyecto_investigacion else "-"] for idx, distincion in enumerate(distinciones, start=1)]
        row = cls._write_table(ws, row, "6.1.- Distinciones y reconocimientos recibidos", ["Nro.", "Fecha", "Descripcion de la distincion", "Proyecto asociado"], distinciones_rows, merge_span=8, date_cols={2})
        row = cls._write_table(ws, row, "6.2.- Participaciones institucionales y academicas relevantes", ["Nro.", "Evento o actividad", "Forma de participacion", "Fecha", "Investigador participante"], participaciones_rows, merge_span=8, date_cols={4})
        visitas_rows = [[idx, visita.razon, visita.procedencia, visita.tipo_visita.nombre if visita.tipo_visita else "-", visita.fecha] for idx, visita in enumerate(visitas, start=1)]
        row = cls._write_table(ws, row, "6.3.- Visitantes del pais y del extranjero", ["Nro.", "Motivo o razon de la visita", "Procedencia", "Tipo de visita", "Fecha"], visitas_rows, merge_span=8, date_cols={5})
        reuniones_grouped = {}
        for trabajo in trabajos_reunion:
            tipo = trabajo.tipo_reunion_cientifica.nombre if trabajo.tipo_reunion_cientifica else "Sin tipo definido"
            reuniones_grouped.setdefault(tipo, []).append([len(reuniones_grouped.get(tipo, [])) + 1, trabajo.titulo_trabajo, trabajo.nombre_reunion, trabajo.procedencia, trabajo.fecha_inicio, cls._join_names(trabajo.investigadores)])
        row = cls._write_grouped_tables(ws, row, "7", "TRABAJOS PRESENTADOS EN CONGRESOS Y REUNIONES CIENTIFICAS CON REFERATO", list(reuniones_grouped.items()), ["Nro.", "Titulo del trabajo", "Reunion cientifica", "Institucion de procedencia", "Fecha de presentacion", "Investigadores participantes"], merge_span=10, date_cols={5})
        articulos_rows = [[idx, articulo.titulo, articulo.descripcion, articulo.fecha_publicacion] for idx, articulo in enumerate(articulos, start=1)]
        row = cls._write_table(ws, row, "8.- TRABAJOS REALIZADOS Y PUBLICADOS", ["Nro.", "Titulo del articulo", "Descripcion o sintesis", "Fecha de publicacion"], articulos_rows, merge_span=8, date_cols={4})
        revistas_rows = [[idx, trabajo.titulo_trabajo, trabajo.nombre_revista, trabajo.editorial, trabajo.issn, trabajo.pais, trabajo.tipo_reunion.nombre if trabajo.tipo_reunion else "-", trabajo.fecha, cls._join_names(trabajo.investigadores)] for idx, trabajo in enumerate(trabajos_revista, start=1)]
        row = cls._write_table(ws, row, "8.1.- Trabajos en revistas con referato", ["Nro.", "Titulo del trabajo", "Revista", "Editorial", "ISSN", "Pais", "Tipo de publicacion", "Fecha", "Investigadores participantes"], revistas_rows, merge_span=10, date_cols={8})
        registros_grouped = {}
        for registro in registros:
            tipo = registro.tipo_registro.nombre if registro.tipo_registro else "Sin tipo definido"
            registros_grouped.setdefault(tipo, []).append([len(registros_grouped.get(tipo, [])) + 1, registro.nombre_articulo, registro.organismo_registrante, registro.fecha_registro])
        row = cls._write_grouped_tables(ws, row, "9", "REGISTROS Y PATENTES", list(registros_grouped.items()), ["Nro.", "Nombre o titulo registrado", "Organismo registrante", "Fecha de registro"], merge_span=8, date_cols={4})

        row = cls._write_section(ws, row, "III.- ACTIVIDADES EN DOCENCIA")
        actividades_rows = [[idx, actividad.curso, actividad.institucion, actividad.investigador.nombre_apellido if actividad.investigador else "-", cls._active_grado_nombre(actividad), actividad.rol_actividad.nombre if actividad.rol_actividad else "-", actividad.fecha_inicio, actividad.fecha_fin] for idx, actividad in enumerate(actividades, start=1)]
        row = cls._write_table(ws, row, "10.- ACTIVIDADES EN DOCENCIA", ["Nro.", "Curso o actividad", "Institucion", "Investigador responsable", "Grado academico vigente", "Rol desempenado", "Fecha de inicio", "Fecha de finalizacion"], actividades_rows, merge_span=10, date_cols={7, 8})

        row = cls._write_section(ws, row, "IV.- VINCULACION CON EL MEDIO SOCIO PRODUCTIVO")
        transferencias_grouped = {}
        total_transferencias = 0.0
        for transferencia in transferencias:
            monto = cls._money(transferencia.monto)
            total_transferencias += monto
            tipo = transferencia.tipo_contrato_transferencia.nombre if transferencia.tipo_contrato_transferencia else "Sin tipo definido"
            adoptantes = ", ".join(sorted({participacion.adoptante.nombre for participacion in transferencia.participaciones if participacion.deleted_at is None and participacion.adoptante is not None})) or "-"
            transferencias_grouped.setdefault(tipo, []).append([len(transferencias_grouped.get(tipo, [])) + 1, transferencia.numero_transferencia, transferencia.denominacion, transferencia.demandante, monto, transferencia.fecha_inicio, transferencia.fecha_fin, adoptantes])
        row = cls._write_grouped_tables(ws, row, "11", "TRANSFERENCIAS SOCIO-PRODUCTIVAS", list(transferencias_grouped.items()), ["Nro.", "Numero de transferencia", "Denominacion", "Demandante", "Monto comprometido", "Fecha de inicio", "Fecha de finalizacion", "Adoptantes vinculados"], accent=True, merge_span=10, date_cols={6, 7}, money_cols={5})
        row = cls._write_totals(ws, row, "Total monto transferencias", [(5, total_transferencias)])

        row = cls._write_section(ws, row, "V.- INFORME SOBRE RENDICION DE CUENTAS")
        erogaciones_grouped = {}
        total_ingresos = 0.0
        total_egresos = 0.0
        for erogacion in erogaciones:
            ingresos = cls._money(erogacion.ingresos)
            egresos = cls._money(erogacion.egresos)
            total_ingresos += ingresos
            total_egresos += egresos
            saldo = ingresos - egresos
            tipo = erogacion.tipo_erogacion.nombre if erogacion.tipo_erogacion else "Sin tipo definido"
            erogaciones_grouped.setdefault(tipo, []).append([len(erogaciones_grouped.get(tipo, [])) + 1, erogacion.numero_erogacion, erogacion.fecha, erogacion.fuente_financiamiento.nombre if erogacion.fuente_financiamiento else "-", ingresos, egresos, saldo])
        row = cls._write_grouped_tables(ws, row, "12", "RESUMEN DE INGRESOS Y EGRESOS (EROGACIONES)", list(erogaciones_grouped.items()), ["Nro.", "Numero de erogacion", "Fecha", "Fuente de financiamiento", "Ingresos", "Egresos", "Saldo resultante"], accent=True, merge_span=9, date_cols={3}, money_cols={5, 6, 7})
        row = cls._write_totals(ws, row, "Totales erogaciones", [(5, total_ingresos), (6, total_egresos), (7, total_ingresos - total_egresos)])

        row = cls._write_section(ws, row, "VI.- PROGRAMA DE ACTIVIDADES FUTURAS")
        planificaciones_rows = [[idx, plan.anio, plan.descripcion] for idx, plan in enumerate(planificaciones, start=1)]
        row = cls._write_table(ws, row, "13.- PLANIFICACIONES FUTURAS DEL GRUPO", ["Nro.", "Anio", "Descripcion"], planificaciones_rows, merge_span=8)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output



